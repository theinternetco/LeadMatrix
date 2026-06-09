import logging
import os
import asyncio
import random
import base64
from datetime import datetime, timezone
from typing import List, Optional
from urllib.parse import quote as url_quote

import io
import requests as _requests
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field, field_validator, model_validator, ConfigDict
from pydantic.alias_generators import to_camel

from app.database import get_db
from app.models import GMBPost, Business, PostHistory
from app.services.post_history import record_post, update_post_status

router = APIRouter(redirect_slashes=False)
logger = logging.getLogger("gmb_posts")

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "")
IMGBB_API_KEY  = os.getenv("IMGBB_API_KEY", "")

ROUTER_VERSION = "2.0"


# ==========================================
# PYDANTIC SCHEMAS
# ==========================================

POST_STATUSES = {"draft", "scheduled", "published", "failed", "pending"}
CTA_ALLOWED   = {"call", "book", "learn_more", "order", "shop", "sign_up", "get_offer"}

DATETIME_FORMATS = [
    "%Y-%m-%dT%H:%M:%S.%fZ",
    "%Y-%m-%dT%H:%M:%SZ",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M",
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%d-%m-%Y %H:%M",
    "%d/%m/%Y %H:%M",
    "%d-%m-%Y %H:%M:%S",
]

CONTENT_ANGLES = [
    "seasonal_offer",
    "service_highlight",
    "before_after",
    "tips_faq",
    "client_result",
    "why_choose_us",
]


def parse_datetime_flexible(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if not isinstance(value, str):
        raise ValueError(f"Cannot parse datetime from type {type(value)}")
    value = value.strip()
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        pass
    for fmt in DATETIME_FORMATS:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    raise ValueError(
        f"Cannot parse '{value}' as datetime. "
        f"Use ISO format: YYYY-MM-DDTHH:MM or DD-MM-YYYY HH:MM"
    )


# ─────────────────────────────────────────
# AutoGenerateRequest
# ─────────────────────────────────────────

class AutoGenerateRequest(BaseModel):
    model_config = ConfigDict(
        alias_generator=to_camel,
        populate_by_name=True,
        extra="ignore",
    )

    business_id:  Optional[int]       = None
    business_ids: Optional[List[int]] = None
    scheduled_at: str = Field(..., description="ISO datetime string for when to publish")

    @model_validator(mode="before")
    @classmethod
    def resolve_business_ids(cls, data):
        if not isinstance(data, dict):
            return data
        bid  = data.get("business_id") or data.get("businessId")
        bids = list(data.get("business_ids") or data.get("businessIds") or [])
        if bid:
            bid = int(bid)
            if bid not in bids:
                bids.insert(0, bid)
        if not bids:
            raise ValueError("business_id or business_ids is required")
        seen, deduped = set(), []
        for x in bids:
            ix = int(x)
            if ix not in seen:
                seen.add(ix)
                deduped.append(ix)
        data["business_ids"] = deduped
        data["business_id"]  = deduped[0]
        return data

    @field_validator("scheduled_at", mode="before")
    @classmethod
    def validate_scheduled_at(cls, v):
        if not v:
            raise ValueError("scheduled_at is required")
        parsed     = parse_datetime_flexible(str(v))
        now        = datetime.now(timezone.utc)
        parsed_utc = parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
        if parsed_utc <= now:
            raise ValueError("scheduled_at must be in the future")
        return v


class AutoGenerateResponse(BaseModel):
    post_id:       int
    business_id:   int
    topic:         str
    description:   str
    image_url:     Optional[str] = None
    scheduled_at:  str
    content_angle: str
    status:        str


class AutoGenerateBulkResponse(BaseModel):
    posts:  List[AutoGenerateResponse]
    total:  int
    errors: List[dict] = []


# ─────────────────────────────────────────
# AutoEditRequest  (NEW)
# ─────────────────────────────────────────

class AutoEditRequest(BaseModel):
    model_config = ConfigDict(
        alias_generator=to_camel,
        populate_by_name=True,
        extra="ignore",
    )

    description:  Optional[str] = Field(None, min_length=1, max_length=1500)
    title:        Optional[str] = Field(None, max_length=200)
    media_url:    Optional[str] = None
    scheduled_at: Optional[str] = None

    @field_validator("media_url", mode="before")
    @classmethod
    def validate_media_url(cls, v):
        if not v:
            return None
        v = str(v).strip()
        if not (v.startswith("http") or v.startswith("data:image/")):
            raise ValueError("media_url must be a valid public URL or base64 data URI")
        return v

    @field_validator("scheduled_at", mode="before")
    @classmethod
    def validate_scheduled_at(cls, v):
        if not v:
            return None
        parsed     = parse_datetime_flexible(str(v))
        now        = datetime.now(timezone.utc)
        parsed_utc = parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
        if parsed_utc <= now:
            raise ValueError("scheduled_at must be in the future")
        return v


# ─────────────────────────────────────────
# GmbPostCreate
# ─────────────────────────────────────────

class GmbPostCreate(BaseModel):
    model_config = ConfigDict(
        alias_generator=to_camel,
        populate_by_name=True,
        extra="ignore",
    )

    business_id:    Optional[int]       = None
    business_ids:   Optional[List[int]] = None
    profile_id:     Optional[str] = None
    content:        Optional[str] = Field(None)
    description:    Optional[str] = Field(None)
    title:          Optional[str] = Field(None, max_length=200)
    media_url:      Optional[str] = None
    post_type:      str = Field("update", pattern="^(update|offer|event)$")
    cta_type:       Optional[str] = None
    cta_value:      Optional[str] = None
    schedule:       bool = False
    scheduled_at:   Optional[str] = None
    scheduled_date: Optional[str] = None

    @model_validator(mode="before")
    @classmethod
    def normalise_fields(cls, data):
        if not isinstance(data, dict):
            return data
        # ── Resolve business_ids (supports both single and multi) ─────────────
        bid  = data.get("business_id") or data.get("businessId")
        bids = list(data.get("business_ids") or data.get("businessIds") or [])
        if bid:
            bid = int(bid)
            if bid not in bids:
                bids.insert(0, bid)
        if not bids:
            raise ValueError("business_id or business_ids is required")
        seen, deduped = set(), []
        for x in bids:
            ix = int(x)
            if ix not in seen:
                seen.add(ix)
                deduped.append(ix)
        data["business_ids"] = deduped
        data["business_id"]  = deduped[0]
        # ── Resolve content fields ────────────────────────────────────────────
        content     = (data.get("content")     or data.get("Content")     or "").strip()
        description = (data.get("description") or data.get("Description") or "").strip()
        resolved    = content or description
        if not resolved:
            raise ValueError("Post content / description is required and cannot be empty.")
        data["content"]     = resolved
        data["description"] = resolved
        sched_at   = (data.get("scheduledAt")   or data.get("scheduled_at")   or "").strip() or None
        sched_date = (data.get("scheduledDate") or data.get("scheduled_date") or "").strip() or None
        canonical  = sched_at or sched_date
        data["scheduled_at"]   = canonical
        data["scheduled_date"] = canonical
        return data

    @field_validator("post_type", mode="before")
    @classmethod
    def normalize_post_type(cls, v):
        return v.lower().strip() if isinstance(v, str) else v

    @field_validator("cta_type", mode="before")
    @classmethod
    def validate_cta_type(cls, v):
        if v and v.lower() not in CTA_ALLOWED:
            raise ValueError(f"cta_type must be one of: {', '.join(sorted(CTA_ALLOWED))}")
        return v.lower().strip() if v else None

    @field_validator("media_url", mode="before")
    @classmethod
    def validate_media_url(cls, v):
        if not v:
            return None
        v = str(v).strip()
        if not (v.startswith("http") or v.startswith("data:image/")):
            raise ValueError("media_url must be a valid public URL (http/https) or a base64 data URI")
        return v

    @model_validator(mode="after")
    def check_schedule_logic(self):
        raw_dt = self.scheduled_at or self.scheduled_date
        if raw_dt:
            try:
                parsed = parse_datetime_flexible(raw_dt)
            except ValueError as e:
                raise ValueError(f"Invalid scheduled_date: {e}")
            now        = datetime.now(timezone.utc)
            parsed_utc = parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
            if parsed_utc <= now:
                raise ValueError("scheduled_date must be in the future")
            self.scheduled_date = raw_dt
            self.scheduled_at   = raw_dt
            self.schedule       = True
        else:
            self.scheduled_date = None
            self.scheduled_at   = None
        if self.schedule and not self.scheduled_date:
            raise ValueError("scheduled_date / scheduledAt is required when schedule=True")
        if not self.schedule:
            self.scheduled_date = None
        if self.cta_type and self.cta_type != "call" and not (self.cta_value or "").strip():
            raise ValueError(
                f"cta_value (URL) is required when cta_type='{self.cta_type}'. "
                "Only cta_type='call' can be used without a URL."
            )
        return self

    def get_description(self) -> str:
        return (self.content or self.description or "").strip()

    def get_parsed_scheduled_date(self) -> Optional[datetime]:
        return parse_datetime_flexible(self.scheduled_date)


# ─────────────────────────────────────────
# GmbPostOut
# ─────────────────────────────────────────

class GmbPostOut(BaseModel):
    id:             int
    business_id:    int
    profile_id:     Optional[str]      = None
    title:          Optional[str]      = None
    description:    str
    content:        str
    media_url:      Optional[str]      = None
    post_type:      str
    cta_type:       Optional[str]      = None
    cta_value:      Optional[str]      = None
    scheduled_date: Optional[datetime] = None
    published_date: Optional[datetime] = None
    status:         str
    error_log:      Optional[str]      = None
    retry_count:    int                = 0
    google_post_id: Optional[str]      = None
    views:          int                = 0
    clicks:         int                = 0
    created_at:     Optional[datetime] = None
    updated_at:     Optional[datetime] = None
    # AI fields
    ai_generated:   bool               = False
    ai_topic:       Optional[str]      = None
    content_angle:  Optional[str]      = None

    model_config = ConfigDict(from_attributes=True)

    @classmethod
    def from_orm_post(cls, post: "GMBPost"):
        _now       = datetime.now(timezone.utc)
        status_val = post.status.value if hasattr(post.status, "value") else (post.status or "draft")
        body       = post.content or ""
        return cls(
            id             = post.id,
            business_id    = post.business_id,
            profile_id     = post.profile_id,
            title          = post.title,
            description    = body,
            content        = body,
            media_url      = post.media_url,
            post_type      = post.post_type or "update",
            cta_type       = post.cta_type,
            cta_value      = post.cta_value,
            scheduled_date = post.scheduled_date,
            published_date = post.published_date,
            status         = status_val,
            error_log      = post.error_log,
            retry_count    = post.retry_count or 0,
            google_post_id = post.google_post_id,
            views          = post.views or 0,
            clicks         = post.clicks or 0,
            created_at     = post.created_at  if post.created_at  is not None else _now,
            updated_at     = post.updated_at  if post.updated_at  is not None else _now,
            ai_generated   = getattr(post, "ai_generated",  False) or False,
            ai_topic       = getattr(post, "ai_topic",       None),
            content_angle  = getattr(post, "content_angle",  None),
        )


# ─────────────────────────────────────────
# GmbPostUpdate
# ─────────────────────────────────────────

class GmbPostUpdate(BaseModel):
    model_config = ConfigDict(
        alias_generator=to_camel,
        populate_by_name=True,
        extra="ignore",
    )

    description:    Optional[str] = Field(None, min_length=1, max_length=1500)
    content:        Optional[str] = Field(None, min_length=1, max_length=1500)
    title:          Optional[str] = Field(None, max_length=200)
    scheduled_date: Optional[str] = None
    cta_type:       Optional[str] = None
    cta_value:      Optional[str] = None
    media_url:      Optional[str] = None

    @model_validator(mode="before")
    @classmethod
    def normalise_content_field(cls, data):
        if isinstance(data, dict):
            content     = data.get("content")
            description = data.get("description")
            resolved    = (content or description or "").strip() or None
            if resolved:
                data["content"]     = resolved
                data["description"] = resolved
        return data

    @field_validator("cta_type", mode="before")
    @classmethod
    def validate_cta_type(cls, v):
        if v and v.lower() not in CTA_ALLOWED:
            raise ValueError(f"cta_type must be one of: {', '.join(sorted(CTA_ALLOWED))}")
        return v.lower().strip() if v else None

    @field_validator("media_url", mode="before")
    @classmethod
    def validate_media_url(cls, v):
        if not v:
            return None
        v = str(v).strip()
        if not (v.startswith("http") or v.startswith("data:image/")):
            raise ValueError("media_url must be a valid public URL or base64 data URI")
        return v

    def get_description(self) -> Optional[str]:
        return (self.content or self.description or "").strip() or None

    def get_parsed_scheduled_date(self) -> Optional[datetime]:
        return parse_datetime_flexible(self.scheduled_date)


# ─────────────────────────────────────────
# BulkConfirmRequest
# ─────────────────────────────────────────

class BulkConfirmRequest(BaseModel):
    post_ids: List[int] = Field(..., min_length=1)


# ==========================================
# HELPERS
# ==========================================

def _get_post_or_404(post_id: int, db: Session) -> GMBPost:
    post = db.query(GMBPost).filter(GMBPost.id == post_id).first()
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    return post


def _resolve_profile_id(post: GMBPost, db: Session) -> str:
    def _is_full_location(val: str) -> bool:
        return bool(val) and "/locations/" in val and not val.startswith("http")

    if post.profile_id and _is_full_location(post.profile_id):
        return post.profile_id
    business = db.query(Business).filter(Business.id == post.business_id).first()
    # Prefer the dedicated gmb_location_id field (API resource path)
    if business and getattr(business, "gmb_location_id", None) and _is_full_location(business.gmb_location_id):
        return business.gmb_location_id
    # Fall back to gmb_url for backwards compat (some businesses may have had API path saved there)
    if business and getattr(business, "gmb_url", None) and _is_full_location(business.gmb_url):
        return business.gmb_url
    biz_display = (getattr(business, "business_name", None) or getattr(business, "name", None) or f"id={post.business_id}") if business else f"id={post.business_id}"
    raise HTTPException(
        status_code=422,
        detail=(
            f"No GMB location set for business '{biz_display}'. "
            "Go to Businesses → Set GMB Location to assign the correct Google Business Profile location."
        ),
    )


def _apply_publish_result(post: GMBPost, result: dict, profile_id: str):
    if result["success"]:
        post.status         = "published"
        post.published_date = datetime.now(timezone.utc)
        post.error_log      = None
        gmb_resp            = result.get("gmb_response") or {}
        post.google_post_id = gmb_resp.get("name") or gmb_resp.get("id")
        if not post.profile_id:
            post.profile_id = profile_id
    else:
        post.status      = "failed"
        post.error_log   = result.get("error")
        post.retry_count = (post.retry_count or 0) + 1


def _pick_content_angle(last_posts: list) -> str:
    """Rotate through content angles based on recent post history."""
    import random
    used = [getattr(p, "content_angle", None) for p in last_posts if getattr(p, "content_angle", None)]
    available = [a for a in CONTENT_ANGLES if a not in used[-3:]]
    if not available:
        available = CONTENT_ANGLES
    return random.choice(available)


def _generate_ai_content(business: Business, angle: str, scheduled_at: datetime) -> dict:
    """
    Generate AI post content using OpenAI/Gemini or fallback to template.
    Returns dict with: topic, description, image_prompt
    """
    biz_name = getattr(business, "business_name", None) or getattr(business, "name", "") or "our business"
    city     = getattr(business, "city",     None) or ""
    category = getattr(business, "category", None) or "local business"
    month    = scheduled_at.strftime("%B")

    angle_templates = {
        "seasonal_offer": {
            "topic": f"{month} Special Offer — {biz_name}",
            "description": (
                f"🌟 {month} Special at {biz_name}! "
                f"We're excited to bring you exclusive deals this season"
                f"{' in ' + city if city else ''}. "
                f"Our team is ready to serve you with the best {category} services. "
                f"Contact us today and take advantage of our limited-time offer. "
                f"Don't miss out — book your appointment now!"
            ),
            "image_prompt": f"Professional {category} service seasonal promotion, bright colors, modern design",
        },
        "service_highlight": {
            "topic": f"Why Choose {biz_name} for {category.title()}",
            "description": (
                f"⭐ Looking for the best {category} service{' in ' + city if city else ''}? "
                f"{biz_name} has you covered! "
                f"Our experienced team delivers exceptional results every time. "
                f"We pride ourselves on quality, reliability, and customer satisfaction. "
                f"Visit us or call today to experience the difference!"
            ),
            "image_prompt": f"Professional {category} team at work, clean modern environment, happy customers",
        },
        "before_after": {
            "topic": f"Real Results from {biz_name}",
            "description": (
                f"🔄 See the transformation! Our clients at {biz_name} love the results. "
                f"We specialize in delivering outstanding {category} outcomes"
                f"{' to ' + city + ' residents' if city else ''}. "
                f"Ready to see your own transformation? "
                f"Book a consultation today and let us show you what we can do!"
            ),
            "image_prompt": f"Before and after transformation showcase, {category} results, professional photography",
        },
        "tips_faq": {
            "topic": f"Pro Tips for {category.title()} — From {biz_name}",
            "description": (
                f"💡 Expert tip from {biz_name}: "
                f"When it comes to {category}, consistency is key! "
                f"Our professionals{' serving ' + city if city else ''} recommend scheduling regular appointments "
                f"for the best long-term results. "
                f"Have questions? We're always here to help — contact us anytime!"
            ),
            "image_prompt": f"Expert tips infographic about {category}, clean design, informative layout",
        },
        "client_result": {
            "topic": f"Client Success Story — {biz_name}",
            "description": (
                f"🏆 Another happy client at {biz_name}! "
                f"We're proud to deliver exceptional {category} results"
                f"{' across ' + city if city else ''}. "
                f"Our commitment to excellence means every client leaves satisfied. "
                f"Join hundreds of happy customers — book your appointment today!"
            ),
            "image_prompt": f"Happy customer testimonial, {category} professional service, warm welcoming atmosphere",
        },
        "why_choose_us": {
            "topic": f"Why {biz_name} Stands Out",
            "description": (
                f"🎯 What makes {biz_name} different? "
                f"✅ Experienced {category} professionals "
                f"✅ Transparent pricing "
                f"✅ Proven results{' in ' + city if city else ''} "
                f"✅ Friendly, dedicated team "
                f"We don't just provide a service — we build lasting relationships. "
                f"Experience the {biz_name} difference today!"
            ),
            "image_prompt": f"Professional {category} business highlights, trust and quality, modern branding",
        },
    }

    template = angle_templates.get(angle, angle_templates["service_highlight"])

    # Try AI generation if available, fall back to template
    try:
        from app.services.ai_content import generate_gmb_post_content
        ai_result = generate_gmb_post_content(
            business_name=biz_name,
            city=city,
            category=category,
            angle=angle,
            month=month,
        )
        if ai_result and ai_result.get("description"):
            template["description"] = ai_result["description"]
            if ai_result.get("topic"):
                template["topic"] = ai_result["topic"]
    except Exception as e:
        logger.info("AI content generation skipped (using template): %s", e)

    return template


def _generate_ai_image(image_prompt: str, business_id: int) -> str:
    """
    Generate AI image via FLUX.1 / Stability AI / or picsum fallback.
    Returns a public URL — NEVER returns None so the preview card always has an image.

    Drop-in upgrade path:
      1. Create app/services/ai_image.py with generate_post_image(prompt, business_id) -> str
      2. This function will automatically use it.
      3. If that module is missing or raises, falls back to picsum (free, no API key needed).
    """
    # ── Try real AI image service ────────────────────────────────────────────
    try:
        from app.services.ai_image import generate_post_image
        url = generate_post_image(prompt=image_prompt, business_id=business_id)
        if url and url.startswith("http"):
            logger.info("AI image generated OK for business_id=%s", business_id)
            return url
    except ImportError:
        logger.info("app.services.ai_image not found — using picsum fallback")
    except Exception as e:
        logger.warning("AI image generation failed (falling back to picsum): %s", e)

    # ── Picsum fallback — always works, seeded by prompt so consistent ───────
    seed = abs(hash(image_prompt + str(business_id))) % 9999
    fallback_url = f"https://picsum.photos/seed/{seed}/800/600"
    logger.info("Using picsum fallback image: %s", fallback_url)
    return fallback_url


# ==========================================
# ⚡ POST /api/gmb-posts/auto-generate
# ==========================================

@router.post("/auto-generate", response_model=AutoGenerateBulkResponse, status_code=201)
def auto_generate_post(payload: AutoGenerateRequest, db: Session = Depends(get_db)):
    """
    AI Auto-Post: Given business_ids + scheduled_at, for EACH business automatically:
    1. Picks the best content angle (rotates 6 types)
    2. Generates SEO-optimized post description
    3. Generates a professional AI image (always returns a URL — never null)
    4. Saves each post with status='pending' (awaiting frontend confirmation)
    Supports single business_id or a list of business_ids.
    """
    business_ids = payload.business_ids or [payload.business_id]

    scheduled_dt  = parse_datetime_flexible(payload.scheduled_at)
    scheduled_utc = scheduled_dt if scheduled_dt.tzinfo else scheduled_dt.replace(tzinfo=timezone.utc)

    results: list[AutoGenerateResponse] = []
    errors:  list[dict] = []

    for bid in business_ids:
        try:
            business = db.query(Business).filter(Business.id == bid).first()
            if not business:
                errors.append({"business_id": bid, "error": f"Business id={bid} not found"})
                continue

            recent_posts = (
                db.query(GMBPost)
                .filter(GMBPost.business_id == bid)
                .order_by(GMBPost.created_at.desc())
                .limit(5)
                .all()
            )
            angle        = _pick_content_angle(recent_posts)
            content_data = _generate_ai_content(business, angle, scheduled_utc)
            topic        = content_data["topic"]
            description  = content_data["description"]
            image_prompt = content_data.get("image_prompt", f"Professional business post for {angle}")
            image_url    = _generate_ai_image(image_prompt, bid)

            post = GMBPost(
                business_id    = bid,
                profile_id     = None,
                title          = topic,
                content        = description,
                media_url      = image_url,
                post_type      = "update",
                cta_type       = None,
                cta_value      = None,
                scheduled_date = scheduled_utc,
                status         = "pending",
                retry_count    = 0,
            )
            if hasattr(post, "ai_generated"):  post.ai_generated  = True
            if hasattr(post, "ai_topic"):      post.ai_topic      = topic
            if hasattr(post, "content_angle"): post.content_angle = angle

            db.add(post)
            db.flush()  # assign post.id without committing

            logger.info("Auto-generate OK: post id=%s business=%s angle=%s", post.id, bid, angle)
            results.append(AutoGenerateResponse(
                post_id       = post.id,
                business_id   = bid,
                topic         = topic,
                description   = description,
                image_url     = image_url,
                scheduled_at  = scheduled_utc.isoformat(),
                content_angle = angle,
                status        = "pending",
            ))
        except Exception as exc:
            logger.error("Auto-generate FAILED for business_id=%s: %s", bid, exc)
            errors.append({"business_id": bid, "error": str(exc)})

    if not results:
        raise HTTPException(
            status_code=422,
            detail=f"All auto-generate attempts failed: {errors}",
        )

    db.commit()
    return AutoGenerateBulkResponse(posts=results, total=len(results), errors=errors)


# ==========================================
# POST /api/gmb-posts/auto-generate/{post_id}/confirm
# ==========================================

@router.post("/auto-generate/{post_id}/confirm")
def confirm_auto_post(post_id: int, db: Session = Depends(get_db)):
    """
    Confirm a pending AI-generated post.
    Changes status: pending → scheduled.
    Called when user clicks 'Confirm & Schedule' in the frontend preview.
    """
    post           = _get_post_or_404(post_id, db)
    current_status = post.status.value if hasattr(post.status, "value") else post.status

    if current_status != "pending":
        raise HTTPException(
            status_code=400,
            detail=f"Only pending posts can be confirmed. Current status: '{current_status}'"
        )
    if not post.scheduled_date:
        raise HTTPException(status_code=400, detail="Post has no scheduled_date set")

    post.status = "scheduled"
    db.commit()
    db.refresh(post)

    record_post(db, post)
    logger.info("Auto-post confirmed: id=%s scheduled=%s", post.id, post.scheduled_date)
    return GmbPostOut.from_orm_post(post)


# ==========================================
# POST /api/gmb-posts/auto-generate/{post_id}/discard  (NEW)
# ==========================================

@router.post("/auto-generate/{post_id}/discard", status_code=200)
def discard_auto_post(post_id: int, db: Session = Depends(get_db)):
    """
    Discard a pending AI-generated post — permanently deletes it.
    Called when user clicks 'Discard' in the frontend preview.
    Only pending posts can be discarded this way.
    """
    post           = _get_post_or_404(post_id, db)
    current_status = post.status.value if hasattr(post.status, "value") else post.status

    if current_status != "pending":
        raise HTTPException(
            status_code=400,
            detail=f"Only pending posts can be discarded. Current status: '{current_status}'"
        )

    db.delete(post)
    db.commit()

    logger.info("Auto-post discarded: id=%s business_id=%s", post_id, post.business_id)
    return {"success": True, "message": f"Post id={post_id} discarded and deleted."}


# ==========================================
# PATCH /api/gmb-posts/auto-generate/{post_id}/edit  (NEW)
# ==========================================

@router.patch("/auto-generate/{post_id}/edit")
def edit_auto_post(post_id: int, payload: AutoEditRequest, db: Session = Depends(get_db)):
    """
    Edit a pending AI-generated post before confirming.
    Allows updating description, title, media_url, and scheduled_at.
    Post stays in 'pending' status — call /confirm to activate.
    Called when user clicks 'Edit' in the frontend preview.
    """
    post           = _get_post_or_404(post_id, db)
    current_status = post.status.value if hasattr(post.status, "value") else post.status

    if current_status != "pending":
        raise HTTPException(
            status_code=400,
            detail=f"Only pending posts can be edited via this endpoint. Current status: '{current_status}'"
        )

    if payload.description is not None:
        post.content = payload.description.strip()
    if payload.title is not None:
        post.title = payload.title.strip()
    if payload.media_url is not None:
        post.media_url = payload.media_url
    if payload.scheduled_at is not None:
        parsed = parse_datetime_flexible(payload.scheduled_at)
        post.scheduled_date = parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)

    db.commit()
    db.refresh(post)

    logger.info("Auto-post edited: id=%s", post.id)
    return GmbPostOut.from_orm_post(post)


# ==========================================
# POST /api/gmb-posts/ — Create (Manual)
# ==========================================

@router.post("", status_code=201)
def create_post(payload: GmbPostCreate, db: Session = Depends(get_db)):
    """
    Create a post for one or multiple businesses.
    Accepts business_id (single) or business_ids (list).
    Returns {"created": N, "posts": [...], "errors": [...]}.
    """
    business_ids     = payload.business_ids or [payload.business_id]
    parsed_scheduled = payload.get_parsed_scheduled_date()
    results: list    = []
    errors:  list    = []

    for bid in business_ids:
        try:
            business = db.query(Business).filter(Business.id == bid).first()
            if not business:
                errors.append({"business_id": bid, "error": f"Business id={bid} not found"})
                continue

            post = GMBPost(
                business_id    = bid,
                profile_id     = payload.profile_id,
                title          = payload.title,
                content        = payload.get_description(),
                media_url      = payload.media_url,
                post_type      = payload.post_type,
                cta_type       = payload.cta_type,
                cta_value      = payload.cta_value,
                scheduled_date = parsed_scheduled,
                status         = "scheduled" if payload.schedule else "draft",
                retry_count    = 0,
            )

            if not payload.schedule:
                db.add(post)
                db.flush()

                profile_id = _resolve_profile_id(post, db)

                from app.services.gmb_publisher import publish_post_to_gmb
                result = publish_post_to_gmb(post, profile_id=profile_id)

                _apply_publish_result(post, result, profile_id)

                if result["success"]:
                    logger.info("Immediate publish OK: post id=%s → GMB: %s", post.id, post.google_post_id)
                else:
                    logger.error("Immediate publish FAILED: post id=%s — %s", post.id, post.error_log)

                db.commit()
                db.refresh(post)

                if post.status == "failed":
                    errors.append({
                        "business_id": bid,
                        "error": f"Post saved (id={post.id}) but GMB publish failed: {post.error_log}",
                    })
                    continue
            else:
                db.add(post)
                db.commit()
                db.refresh(post)
                logger.info("Post id=%s scheduled for %s (business_id=%s)", post.id, post.scheduled_date, bid)

            biz_name_str = getattr(business, "business_name", None) or getattr(business, "name", None) or f"Business #{bid}"
            record_post(db, post, business_name=biz_name_str)
            results.append(GmbPostOut.from_orm_post(post).model_dump())

        except HTTPException as he:
            errors.append({"business_id": bid, "error": he.detail})
        except Exception as exc:
            logger.error("create_post FAILED for business_id=%s: %s", bid, exc)
            errors.append({"business_id": bid, "error": str(exc)})

    if not results and errors:
        first_err = errors[0]["error"] if len(errors) == 1 else f"All {len(errors)} post creation(s) failed"
        raise HTTPException(status_code=422, detail=first_err)

    return {"created": len(results), "posts": results, "errors": errors}


# ==========================================
# POST /api/gmb-posts/bulk-confirm
# ==========================================

@router.post("/bulk-confirm")
def bulk_confirm_posts(payload: BulkConfirmRequest, db: Session = Depends(get_db)):
    """
    Confirm multiple pending AI-generated posts in one call.
    Changes status: pending → scheduled for each post_id.
    Failures for individual posts are returned in errors list without aborting others.
    """
    confirmed: list[int] = []
    errors:    list[dict] = []

    for post_id in payload.post_ids:
        try:
            post           = db.query(GMBPost).filter(GMBPost.id == post_id).first()
            if not post:
                errors.append({"post_id": post_id, "error": "Post not found"})
                continue
            current_status = post.status.value if hasattr(post.status, "value") else post.status
            if current_status != "pending":
                errors.append({"post_id": post_id, "error": f"Post is not pending (status: {current_status})"})
                continue
            if not post.scheduled_date:
                errors.append({"post_id": post_id, "error": "Post has no scheduled_date set"})
                continue
            post.status = "scheduled"
            confirmed.append(post_id)
        except Exception as exc:
            errors.append({"post_id": post_id, "error": str(exc)})

    db.commit()
    logger.info("Bulk confirm: %d confirmed, %d errors", len(confirmed), len(errors))
    return {"confirmed": len(confirmed), "post_ids": confirmed, "errors": errors}


# ==========================================
# GET /api/gmb-posts/history/export — Monthly Excel Export
# ==========================================

@router.get("/history/export")
def export_post_history(
    month: Optional[int] = Query(None, ge=1, le=12, description="Month to export (default: current month)"),
    year:  Optional[int] = Query(None, ge=2020,     description="Year to export (default: current year)"),
    db: Session = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    now           = datetime.now(timezone.utc)
    export_month  = month or now.month
    export_year   = year  or now.year

    rows = (
        db.query(PostHistory)
        .filter(PostHistory.month == export_month, PostHistory.year == export_year)
        .order_by(PostHistory.business_name.asc(), PostHistory.scheduled_date.asc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    month_name = datetime(export_year, export_month, 1).strftime("%B %Y")
    ws.title = month_name[:31]

    # ── Header ─────────────────────────────────────────────────────────────────
    headers = ["Client / Business", "Title", "Type", "Status", "Scheduled Date", "Published Date"]
    header_fill = PatternFill("solid", fgColor="6366F1")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22

    # ── Data rows, colour-coded by status ──────────────────────────────────────
    STATUS_COLORS = {
        "published": "D1FAE5",  # green tint
        "scheduled": "DBEAFE",  # blue tint
        "failed":    "FEE2E2",  # red tint
    }
    current_biz   = None
    biz_fill_alt  = PatternFill("solid", fgColor="F8F9FF")
    biz_fill_main = PatternFill("solid", fgColor="FFFFFF")
    biz_toggle    = False

    for row_idx, r in enumerate(rows, 2):
        if r.business_name != current_biz:
            current_biz = r.business_name
            biz_toggle  = not biz_toggle

        status_str = r.status or ""
        row_fill   = PatternFill("solid", fgColor=STATUS_COLORS.get(status_str, "F1F5F9"))

        def _fmt_dt(dt):
            if not dt:
                return ""
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.strftime("%d %b %Y %H:%M")

        values = [
            r.business_name,
            (r.content or "")[:100],
            r.post_type or "",
            status_str.upper(),
            _fmt_dt(r.scheduled_date),
            _fmt_dt(r.published_date),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(wrap_text=(col == 3), vertical="top")

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [28, 50, 12, 14, 22, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Stream response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"posts_{export_year}_{export_month:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ==========================================
# GET /api/gmb-posts/ — List (with pagination metadata)
# ==========================================

@router.get("")
def list_posts(
    business_id: Optional[int] = Query(None),
    profile_id:  Optional[str] = Query(None),
    status:      Optional[str] = Query(None, description="Single status or comma-separated: pending,scheduled"),
    limit:       int = Query(50, ge=1, le=200),
    offset:      int = Query(0, ge=0),
    db: Session = Depends(get_db),
):
    q = db.query(GMBPost)

    if business_id:
        q = q.filter(GMBPost.business_id == business_id)
    if profile_id:
        q = q.filter(GMBPost.profile_id == profile_id)

    if status:
        # Support comma-separated: ?status=pending,scheduled
        requested = [s.strip() for s in status.split(",") if s.strip()]
        invalid   = [s for s in requested if s not in POST_STATUSES]
        if invalid:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid status value(s): {invalid}. Must be one of: {', '.join(sorted(POST_STATUSES))}"
            )
        if len(requested) == 1:
            q = q.filter(GMBPost.status == requested[0])
        else:
            q = q.filter(GMBPost.status.in_(requested))

    total = q.count()
    posts = (
        q.order_by(GMBPost.created_at.desc().nullslast())
         .offset(offset)
         .limit(limit)
         .all()
    )

    return {
        "total":   total,
        "limit":   limit,
        "offset":  offset,
        "has_more": (offset + limit) < total,
        "items":   [GmbPostOut.from_orm_post(p) for p in posts],
    }


# ==========================================
# GET /api/gmb-posts/stats/summary
# ==========================================

@router.get("/stats/summary")
def posts_summary(
    business_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(GMBPost)
    if business_id:
        q = q.filter(GMBPost.business_id == business_id)
    posts = q.all()

    def _s(p):
        return p.status.value if hasattr(p.status, "value") else (p.status or "draft")

    return {
        "total":        len(posts),
        "draft":        sum(1 for p in posts if _s(p) == "draft"),
        "scheduled":    sum(1 for p in posts if _s(p) == "scheduled"),
        "published":    sum(1 for p in posts if _s(p) == "published"),
        "failed":       sum(1 for p in posts if _s(p) == "failed"),
        "pending":      sum(1 for p in posts if _s(p) == "pending"),
        "ai_generated": sum(1 for p in posts if getattr(p, "ai_generated", False)),
    }


# ==========================================
# GET /api/gmb-posts/{id} — Single
# ==========================================

@router.get("/{post_id}")
def get_post(post_id: int, db: Session = Depends(get_db)):
    return GmbPostOut.from_orm_post(_get_post_or_404(post_id, db))


# ==========================================
# PATCH /api/gmb-posts/{id} — Edit (Manual)
# ==========================================

@router.patch("/{post_id}")
def update_post(post_id: int, payload: GmbPostUpdate, db: Session = Depends(get_db)):
    post           = _get_post_or_404(post_id, db)
    current_status = post.status.value if hasattr(post.status, "value") else post.status

    if current_status == "published":
        raise HTTPException(status_code=400, detail="Cannot edit an already-published post")

    body = payload.get_description()
    if body:
        post.content = body
    if payload.title is not None:
        post.title = payload.title.strip()
    if payload.media_url is not None:
        post.media_url = payload.media_url
    if payload.cta_type is not None:
        post.cta_type = payload.cta_type
    if payload.cta_value is not None:
        post.cta_value = payload.cta_value
    if payload.scheduled_date is not None:
        try:
            parsed = payload.get_parsed_scheduled_date()
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
        now        = datetime.now(timezone.utc)
        parsed_utc = parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
        if parsed_utc <= now:
            raise HTTPException(status_code=400, detail="scheduled_date must be in the future")
        post.scheduled_date = parsed
        post.status         = "scheduled"

    db.commit()
    db.refresh(post)
    return GmbPostOut.from_orm_post(post)


# ==========================================
# DELETE /api/gmb-posts/{id}
# ==========================================

@router.delete("/{post_id}", status_code=204)
def delete_post(post_id: int, db: Session = Depends(get_db)):
    post           = _get_post_or_404(post_id, db)
    current_status = post.status.value if hasattr(post.status, "value") else post.status
    if current_status == "published":
        raise HTTPException(status_code=400, detail="Cannot delete a published post")
    db.delete(post)
    db.commit()


# ==========================================
# POST /api/gmb-posts/{id}/trigger — Force publish
# ==========================================

@router.post("/{post_id}/trigger")
def trigger_post_now(post_id: int, db: Session = Depends(get_db)):
    post           = _get_post_or_404(post_id, db)
    current_status = post.status.value if hasattr(post.status, "value") else post.status

    if current_status == "published":
        raise HTTPException(status_code=400, detail="Post is already published")

    profile_id = _resolve_profile_id(post, db)
    logger.info("Trigger post id=%s using profile_id=%s", post.id, profile_id)

    from app.services.gmb_publisher import publish_post_to_gmb
    result = publish_post_to_gmb(post, profile_id=profile_id)
    _apply_publish_result(post, result, profile_id)

    if result["success"]:
        logger.info("Manual trigger OK: post id=%s → GMB: %s", post.id, post.google_post_id)
    else:
        logger.error("Manual trigger FAILED: post id=%s — %s", post.id, post.error_log)

    db.commit()
    db.refresh(post)

    if post.status == "failed":
        raise HTTPException(
            status_code=422,
            detail=f"GMB publish failed [profile={profile_id}]: {post.error_log}",
        )

    return GmbPostOut.from_orm_post(post)


# ==========================================
# POST /api/gmb-posts/{id}/retry
# ==========================================

@router.post("/{post_id}/retry")
def retry_failed_post(post_id: int, db: Session = Depends(get_db)):
    post           = _get_post_or_404(post_id, db)
    current_status = post.status.value if hasattr(post.status, "value") else post.status

    if current_status != "failed":
        raise HTTPException(status_code=400, detail="Only failed posts can be retried")

    profile_id = _resolve_profile_id(post, db)

    from app.services.gmb_publisher import publish_post_to_gmb
    result = publish_post_to_gmb(post, profile_id=profile_id)
    _apply_publish_result(post, result, profile_id)

    if result["success"]:
        logger.info("Retry OK: post id=%s → GMB: %s", post.id, post.google_post_id)
    else:
        logger.error("Retry FAILED: post id=%s — %s", post.id, post.error_log)

    db.commit()
    db.refresh(post)
    return GmbPostOut.from_orm_post(post)


# ==========================================
# POST /api/gmb-posts/{id}/reschedule
# ==========================================

@router.post("/{post_id}/reschedule")
def reschedule_post(
    post_id:  int,
    new_time: str,
    db: Session = Depends(get_db),
):
    post           = _get_post_or_404(post_id, db)
    current_status = post.status.value if hasattr(post.status, "value") else post.status

    if current_status == "published":
        raise HTTPException(status_code=400, detail="Cannot reschedule a published post")

    try:
        parsed = parse_datetime_flexible(new_time)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    now    = datetime.now(timezone.utc)
    nt_utc = parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    if nt_utc <= now:
        raise HTTPException(status_code=400, detail="new_time must be in the future")

    post.scheduled_date = parsed
    post.status         = "scheduled"
    post.error_log      = None

    db.commit()
    db.refresh(post)
    return GmbPostOut.from_orm_post(post)


# ==========================================
# AI AUTO-POST ENDPOINTS
# ==========================================

class AIGenerateRequest(BaseModel):
    topic:       str
    business_id: int


class AIRegenerateImageRequest(BaseModel):
    topic:       str
    business_id: int


class AIPostRequest(BaseModel):
    business_id:  int
    text:         str
    image_url:    str
    topic:        str
    scheduled_at: Optional[str] = None  # ISO string → scheduled; None → draft


def _call_gemini(prompt: str) -> str:
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY not set")
    import time
    # Try primary model first, fall back to flash-lite if 429
    models = ["gemini-2.0-flash", "gemini-2.0-flash-lite", "gemini-1.5-flash-latest"]
    base   = "https://generativelanguage.googleapis.com/v1beta/models"
    delays = [5, 15]
    last_err = None
    for model in models:
        url = f"{base}/{model}:generateContent?key={GEMINI_API_KEY}"
        for attempt, delay in enumerate(delays + [None], start=1):
            resp = _requests.post(
                url,
                headers={"Content-Type": "application/json"},
                json={"contents": [{"parts": [{"text": prompt}]}]},
                timeout=40,
            )
            if resp.status_code == 429:
                try:
                    body = resp.json()
                    reason = body.get("error", {}).get("message", "quota exceeded")
                except Exception:
                    reason = resp.text[:200]
                logger.warning("Gemini 429 on %s (attempt %d): %s", model, attempt, reason)
                if delay is None:
                    last_err = f"Rate limit on {model}: {reason}"
                    break  # try next model
                time.sleep(delay)
                continue
            resp.raise_for_status()
            return resp.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
    raise RuntimeError(last_err or "All Gemini models returned 429 — quota exhausted")


def _generate_post_text(topic: str, name: str, city: str, category: str, keywords: list) -> str:
    kw_str = ", ".join(keywords) if keywords else ""
    kw_line = f"Naturally include these keywords if relevant: {kw_str}" if kw_str else ""
    prompt = (
        f"Write a Google Business post for {name}, a {category or 'local'} business"
        + (f" in {city}" if city else "")
        + f".\n\nTopic: {topic}\n{kw_line}\n\n"
        "Rules:\n"
        "- Under 1500 characters\n"
        "- Friendly, professional tone\n"
        "- End with a short call to action\n"
        "- No hashtags\n"
        "- Output the post text only, nothing else"
    )
    return _call_gemini(prompt)


def _generate_and_upload_image(topic: str, name: str, category: str) -> str:
    image_prompt = (
        f"Professional high-quality photo for a {category or 'local'} business, "
        f"{topic}, clean modern style, square 1:1 format, bright and inviting"
    )
    encoded = url_quote(image_prompt)
    seed    = random.randint(1, 999999)
    poll_url = f"https://image.pollinations.ai/prompt/{encoded}?width=1024&height=1024&nologo=true&model=flux&seed={seed}"

    img_resp = _requests.get(poll_url, timeout=90)
    img_resp.raise_for_status()

    if not IMGBB_API_KEY:
        raise RuntimeError("IMGBB_API_KEY not set — cannot upload generated image")

    img_b64 = base64.b64encode(img_resp.content).decode("utf-8")
    up_resp = _requests.post(
        "https://api.imgbb.com/1/upload",
        data={"key": IMGBB_API_KEY, "image": img_b64, "expiration": 2592000},
        timeout=30,
    )
    up_resp.raise_for_status()
    return up_resp.json()["data"]["url"]


@router.post("/ai-generate")
async def ai_generate(req: AIGenerateRequest, db: Session = Depends(get_db)):
    if not GEMINI_API_KEY:
        raise HTTPException(status_code=503, detail="GEMINI_API_KEY not configured")

    business = db.query(Business).filter(Business.id == req.business_id).first()
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")

    name     = business.business_name or business.name or ""
    city     = business.city     or ""
    category = business.category or ""
    keywords = business.keywords or []

    try:
        text = await asyncio.to_thread(_generate_post_text, req.topic, name, city, category, keywords)
    except Exception as e:
        logger.error("AI text generation failed: %s", e)
        raise HTTPException(status_code=502, detail=f"Text generation failed: {e}")

    try:
        image_url = await asyncio.to_thread(_generate_and_upload_image, req.topic, name, category)
    except Exception as e:
        logger.error("AI image generation failed: %s", e)
        raise HTTPException(status_code=502, detail=f"Image generation failed: {e}")

    return {"text": text, "image_url": image_url}


@router.post("/ai-regenerate-image")
async def ai_regenerate_image(req: AIRegenerateImageRequest, db: Session = Depends(get_db)):
    business = db.query(Business).filter(Business.id == req.business_id).first()
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")

    name     = business.business_name or business.name or ""
    category = business.category or ""

    try:
        image_url = await asyncio.to_thread(_generate_and_upload_image, req.topic, name, category)
    except Exception as e:
        logger.error("AI image regeneration failed: %s", e)
        raise HTTPException(status_code=502, detail=f"Image generation failed: {e}")

    return {"image_url": image_url}


@router.post("/ai-post")
def ai_post(req: AIPostRequest, db: Session = Depends(get_db)):
    business = db.query(Business).filter(Business.id == req.business_id).first()
    if not business:
        raise HTTPException(status_code=404, detail="Business not found")

    scheduled_dt = None
    status       = "draft"
    if req.scheduled_at:
        scheduled_dt = parse_datetime_flexible(req.scheduled_at)
        status       = "scheduled"

    post = GMBPost(
        business_id   = req.business_id,
        description   = req.text,
        media_url     = req.image_url or None,
        post_type     = "update",
        status        = status,
        scheduled_date = scheduled_dt,
        ai_generated  = True,
        ai_topic      = req.topic[:300] if req.topic else None,
        profile_id    = business.gmb_url,
    )
    db.add(post)
    db.commit()
    db.refresh(post)

    try:
        record_post(db, post, "ai_created")
    except Exception:
        pass

    return {"success": True, "post_id": post.id, "status": status}